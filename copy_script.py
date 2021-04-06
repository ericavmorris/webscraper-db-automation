import csv
from openpyxl import load_workbook
import os


project_dir = os.getcwd()
myfile=os.path.join(project_dir, 'Covid19_Dashboard.xlsx')

def copy_database_into_dashboard(database_file, dashboard_file):

    workbook = load_workbook(dashboard_file)
    input_sheet = workbook['Input']
    worksheet_names = workbook.sheetnames
    sheet_index = worksheet_names.index('Input')
    workbook.active = sheet_index

    with open(database_file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        #step1:
        column_number = 5
        for row in csv_reader:
            row_number = 1
            for value in row:
                input_sheet.cell(column=column_number, row=row_number, value=value)
                column_number += 1
            column_number += 1

    workbook.save(filename = myfile)
