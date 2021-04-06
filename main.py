import csv
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import shutil, os, sys, fnmatch
import pdb
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import TimeoutException

from company_name_mapping import company_mappings
from functions import *
import collections

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

driver = webdriver.Chrome(ChromeDriverManager().install())

def sign_in(driver):
    input_fields = driver.find_elements_by_class_name('textInput')
    print(input_fields[0])
    print(input_fields[1])
    username_input_field = input_fields[0]
    password_input_field = input_fields[1]
    username_input_field.send_keys("username")
    password_input_field.send_keys("password")
    sign_in_button = driver.find_element_by_class_name("signinButton")
    sign_in_button.click()

def select_db(driver, db_name):
    searchbar_select_database = driver.find_element_by_id("communitySetsComponent")
    searchbar = driver.find_element_by_xpath("//input[@placeholder='Search Databases']")
    searchbar.send_keys(db_name)
    searchbar.send_keys(Keys.ENTER)
    company_listing = driver.find_element_by_tag_name('td')
    company_listing.click()
    time.sleep(.5)
    company_listing.click()
    time.sleep(.5)
    submit_button = driver.find_element_by_id("submitButton")
    submit_button.click()
    submit_button.click()
#pdb.set_trace()

#"Interactive Reports: New Button"
def click_extract_btn(driver):
    scheduled_extract_button = driver.find_element_by_xpath("//img[@alt='Interactive Reports: New']")
    scheduled_extract_button.click()

def filter_for_db(var):
    class_name = "v-table-cell-content v-table-cell-content-truncate"
    if (var.get_attribute("class") == class_name):
        return True
    else:
        return False

def filter_for_dl_btn(var):
    class_name = "v-button-wrap"

    if (var.get_attribute("class") == class_name):
        return True
    else:
        return False

def grab_most_recent_db(driver):
    td_tags_array = driver.find_elements_by_tag_name('td')
    filtered_td_tags = list(filter(filter_for_db, td_tags_array))

    report = filtered_td_tags[1]
    report.click()
    report.click()


def click_dl_btn(driver):
    span_tags = driver.find_elements_by_tag_name('span')
    filtered_span_tags = list(filter(filter_for_dl_btn, span_tags))
    download_button = filtered_span_tags[4]
    ActionChains(driver).move_to_element_with_offset(download_button,0,10).click().perform()
    time.sleep(1.5)


def find_file(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
    if len(result) > 0:
        return result[0]
    else:
        return False

def delete_files_in_directory(pattern, path):
    result = []
    for root, dirs, files in os.walk(path):
        for name in files:
            if fnmatch.fnmatch(name, pattern):
                result.append(os.path.join(root, name))
        break
    for file in result:
        os.remove(file)


def create_company_directory(company_name):
    if not os.path.exists(company_mappings[company_name]):
        os.makedirs(company_mappings[company_name])

def is_db_in_folder(expected_destination_folder_path):

    present_csv_file = find_file('*.csv', expected_destination_folder_path)

    if not present_csv_file:
        return False
    else:
        return True

def diff_lists(first, second):
        second = set(second)
        return [item for item in first if item not in second]


def copy_database_into_dashboard(database_file, dashboard_file):

    workbook = load_workbook(dashboard_file)
    input_sheet = workbook['Input']
    worksheet_names = workbook.sheetnames
    sheet_index = worksheet_names.index('Input')
    workbook.active = sheet_index

    with open(database_file) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        #step1:
        column_number = 5
        for row in csv_reader:
            row_number = 1
            for value in row:
                input_sheet.cell(column=column_number, row=row_number, value=value)
                column_number += 1
            column_number += 1

    workbook.save(filename = dashboard_file)
#

def config_and_navigate_to_login_page():

    is_windows = sys.platform.startswith('win')

    if is_windows:
        chrome_driver_executable = 'chromedriver.exe'
    else:
        chrome_driver_executable = 'chromedriver'

    print(os.path.join(project_dir, chrome_driver_executable))
    # pdb.set_trace()

#    driver = webdriver.Chrome(executable_path=os.path.join(project_dir, chrome_driver_executable), options=options)
#    driver.implicitly_wait(10)
    driver.get("your database")

    return driver

##############################################################
##############################################################
##############################################################
##############################################################
##############################################################
##############################################################
##############################################################

#passing driver object and target databases list we want to DL
#def download_csvs(driver, target_dbs):
    dashboard_file=os.path.join(project_dir, 'Covid19_Dashboard.xlsx')

    #step 1: print database list
    print("{}".format(target_dbs))
    #step 2: create "finished" boolean
    finished = False
    #step 3: create empty downloaded db array to be filled w/ the names of companies for which db has been dl
    downloaded_databases = []
    while not finished:
    #while value is false, keep looping until it becomes true
        try:
            for db in target_dbs:
                time.sleep(1)
                if db not in downloaded_databases:
                    #create folder with company name
                    create_company_directory(db)
                    #os.path.join : allows to pass different strings and make them one path
                    destination_folder = os.path.join(project_dir, company_mappings[db])
                    time.sleep(1)
                    # In browser, enter company name in search bar, and click select button
                    select_db(driver, db)
                    time.sleep(1)
                    # In browser, cick the extract btn to redirect to database list page
                    click_extract_btn(driver)
                    # In browser, select the most recent database
                    grab_most_recent_db(driver)
                    # In browser, click the download button
                    click_dl_btn(driver)
                    time.sleep(1)
                    # grab the newly downloaded csv
                    downloaded_csv = find_file('*.csv', project_dir)
                    # copy it into the company's folder
                    delete_files_in_directory('*.csv', destination_folder)
                    shutil.copy(downloaded_csv, destination_folder)
                    # delete it from the project's directory
                    #Copy dashboard excel file into destination_folder
                    shutil.copy(dashboard_file, destination_folder)
                    # At this point, we have our db csv in project dir, and a xlsx in destination folder
                    target_dashboard_path = find_file('*.xlsx', destination_folder)
                    copy_database_into_dashboard(downloaded_csv, target_dashboard_path)
                    time.sleep(1)

                    delete_files_in_directory('*.csv', project_dir)
                    #append: adds company name to downloaded db array
                    downloaded_databases.append(db)

                    # go back to database selection page
                    driver.get("your database")
        # TODO: Add checker if target and downloaded are equal
            if collections.Counter(target_dbs) == collections.Counter(downloaded_databases):
                finished = True
        except:
            driver.quit()
            driver = config_and_navigate_to_login_page()
            sign_in(driver)
            continue
    driver.quit()
    print("================== DONE ========================")
    print("================== DONE ========================")
    print("================== DONE ========================")
    print("================== DONE ========================")


options = webdriver.ChromeOptions()
project_dir = os.getcwd()
preferences = {
                    "download.default_directory": r"{}".format(project_dir),
                    "directory_upgrade": True
                }
options.add_experimental_option("prefs", preferences)

driver = config_and_navigate_to_login_page()
sign_in(driver)


#target_databases = sorted(company_mappings.keys())

#download_csvs(driver, target_databases)




# ......................next step below
# database_dropdown_btn = driver.find_element_by_xpath("//span[@class='v-button-wrap']")
# print(database_dropdown_btn.location)
# ActionChains(driver).move_to_element_with_offset(database_dropdown_btn,0,10).click().perform()
#
# searchbar2 = driver.find_element_by_xpath("//input[@placeholder='Search Databases']")
# searchbar2.send_keys("Accurate Metal Machining Inc")
# searchbar2.send_keys(Keys.ENTER)


    # next_company_listing = driver.find_element_by_xpath("//*[contains(text(), 'Accurate Metal Machining Inc')]")
    # next_company_listing.click()
    #
    # next_company_listing = driver.find_element_by_xpath("//*[contains(text(), 'Accurate Metal Machining Inc')]")
    # next_company_listing.click()


#
# company_title_placeholder = list(span_tags)[3]
# print(company_title_placeholder.text)

# company_title_placeholder = WebDriverWait(driver, 10).until(
#         EC.text_to_be_present_in_element_value((By.XPATH, "//span[contains(text(), 'Accurate Metal Machining Inc')"))
#     )


# company_title_placeholder = WebDriverWait(driver, 20).until(EC.text_to_be_present_in_element((By.XPATH("//span[@class='v-button-caption']")), 'Accurate Metal Machining Inc'))
#
# td_tags_array = driver.find_elements_by_tag_name('td')
#
# filtered_td_tags = list(filter(filter_for_db, td_tags_array))
#
# report = filtered_td_tags[1]
# report.click()
# report.click()
#
# span_tags = driver.find_elements_by_tag_name('span')
#
# filtered_span_tags = list(filter(filter_for_dl_btn, span_tags))
#
# download_button = filtered_span_tags[4]
#
# ActionChains(driver).move_to_element_with_offset(download_button,0,10).click().perform()

# *************************************
# *************************************
# *************************************
# *************************************
# *************************************
# *************************************

# THIS IS GOING TO BE NEEDED LATER
# database_dropdown_btn = driver.find_element_by_xpath("//span[@class='v-button-wrap']")
# print(database_dropdown_btn.location)
# ActionChains(driver).move_to_element_with_offset(database_dropdown_btn,0,10).click().perform()
#
#
# span_tags = driver.find_elements_by_tag_name('span')
# for span in span_tags:
#     print('span is: , {}'.format(span.get_attribute("class")))
#



#
# def select_db():
#     select_btn_clicked = False
#     while not select_btn_clicked:
#         # next_company_listing = driver.find_element_by_xpath("//*[contains(text(), 'Accurate Metal Machining Inc')]")
#         next_company_listing = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Accurate Metal Machining Inc')]")))
#
#         next_company_listing.click()
#         next_company_listing = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Accurate Metal Machining Inc')]")))
#
#         # next_company_listing = driver.find_element_by_xpath("//*[contains(text(), 'Accurate Metal Machining Inc')]")
#         # next_company_listing.click()
#
#         # click once, then check if error box popped up
#         # if it did, press esc else click the listing again and then click on select btn
#         try:
#             alert = driver.find_element_by_xpath("//div[role='alert']")
#             print(alert.get_attribute("class"))
#             if bool(alert):
#                 print("alert")
#                 ActionChains(driver).send_keys(Keys.ESCAPE).perform()
#             else:
#                 db_select_btn = driver.find_element_by_id('select')
#                 print('db_select_btn is: , {} | class: {}'.format(db_select_btn.text, db_select_btn.get_attribute("id")))
#                 ActionChains(driver).move_to_element_with_offset(db_select_btn,0,5).click().perform()
#                 # db_select_btn = driver.find_elements_by_class_name('v-slot v-slot-nx-action-button')
#                 # print(db_select_btn).get_attribute(class_name)
#                 db_select_btn.click()
#                 db_select_btn.click()
#                 select_btn_clicked = True
#         except:
#             db_select_btn = driver.find_element_by_id('select')
#             print('db_select_btn is: , {} | class: {}'.format(db_select_btn.text, db_select_btn.get_attribute("id")))
#             ActionChains(driver).move_to_element_with_offset(db_select_btn,0,5).click().perform()
#             # db_select_btn = driver.find_elements_by_class_name('v-slot v-slot-nx-action-button')
#             # print(db_select_btn).get_attribute(class_name)
#             db_select_btn.click()
#             db_select_btn.click()
#             select_btn_clicked = True
#
#
# select_db()
